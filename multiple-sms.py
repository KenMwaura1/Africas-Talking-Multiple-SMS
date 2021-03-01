import os
import africastalking as at
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

username = os.getenv("username")
api_key = os.getenv("api_key")
at.initialize(username, api_key)
sms = at.SMS

at.initialize(username, api_key)
wb = load_workbook('sample.xlsx')
print(wb.sheetnames)
sheet1 = wb['Sheet1']
# print(sheet1['B2'].value)
names_cell_range = sheet1['B2:B4']
number_cell_range = sheet1['C2:C4']


def number_generator(cell_range):
    for number in number_cell_range:
        for n in number:
            yield n.value


def name_generator():
    for name in names_cell_range:
        for n in name:
            yield n.value


def send_messages():
    client_name = ''
    recipients = ''
    for each in name_generator():
        client_name = each
        for n in number_generator(number_cell_range):
            recipients = str(f"+254{n}" )
            print(recipients)

            message = f"hey from python {client_name}"
            try:
                response = sms.send(message,[recipients,])
                print(response)
            except Exception as e:
                print(f"Uh oh we have a problem: {e}")


send_messages()
