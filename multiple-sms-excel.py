import os
import africastalking as at
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

username = os.getenv("username")
api_key = os.getenv("api_key")
at.initialize(username, api_key)
sms = at.SMS

at.initialize(username, api_key)
wb = load_workbook('sample.xlsx')
print(wb.sheetnames)
sheet1 = wb['Sheet1']

names_cell_range = sheet1['B2:B4']
number_cell_range = sheet1['C2:C4']


def send_messages():
    for row in sheet1.iter_rows(values_only=True):
        name = row[1]
        number = f"+254{row[2]}"
        print(name,number)
        message = f"hey {name} from python using africas talking API"
        try:
            response = sms.send(message, [number])
            print(response)
        except Exception as e:
            print(f"Uh oh we have a problem: {e}")


send_messages()

